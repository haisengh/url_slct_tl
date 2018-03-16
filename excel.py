from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")
print(wb.sheetnames)
print(ws.title)